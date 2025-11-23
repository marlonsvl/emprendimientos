import openpyxl
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.core.exceptions import ValidationError
from django.db import connection
from gastronomia.models import Establecimiento  


class Command(BaseCommand):
    help = 'Load Establecimiento data from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Path to the Excel file to import'
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Clear existing data before importing'
        )
        parser.add_argument(
            '--batch-size',
            type=int,
            default=500,
            help='Batch size for bulk create (default: 500)'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='Sheet1',
            help='Name of the sheet to import (default: Sheet1)'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        batch_size = options['batch_size']
        sheet_name = options['sheet']
        
        if options['clear']:
            Establecimiento.objects.all().delete()
            # Reset the auto-increment counter to 1
            with connection.cursor() as cursor:
                # For PostgreSQL
                try:
                    cursor.execute("ALTER SEQUENCE gastronomia_establecimiento_id_seq RESTART WITH 1;")
                except:
                    pass  # For other databases that don't use this syntax
            self.stdout.write(self.style.WARNING('Cleared existing data and reset ID counter'))

        try:
            # Load the Excel workbook
            wb = openpyxl.load_workbook(excel_file)
            
            # Get the specified sheet
            if sheet_name not in wb.sheetnames:
                self.stdout.write(
                    self.style.ERROR(f'Sheet "{sheet_name}" not found. Available sheets: {wb.sheetnames}')
                )
                return
            
            ws = wb[sheet_name]
            
            # Get headers from the first row
            headers = [cell.value for cell in ws[1]]
            headers = [h.strip() if h else None for h in headers]
            
            self.stdout.write(self.style.WARNING(f'Excel columns: {headers}'))
            
            created_count = 0
            error_count = 0
            batch = []
            
            # Iterate through rows starting from row 2 (skip header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Create dictionary from headers and row values
                    row_dict = {}
                    for header, value in zip(headers, row):
                        if header:
                            row_dict[header] = value
                    
                    # Convert empty strings to appropriate defaults
                    data = self._clean_row(row_dict)
                    
                    # Create the Establecimiento instance
                    establecimiento = Establecimiento(**data)
                    batch.append(establecimiento)
                    
                    # Bulk create when batch reaches batch_size
                    if len(batch) >= batch_size:
                        Establecimiento.objects.bulk_create(batch, ignore_conflicts=False)
                        created_count += len(batch)
                        self.stdout.write(f'Created {created_count} records so far...')
                        batch = []
                        
                        # Clear database connections to prevent memory issues
                        connection.close()
                    
                except Exception as e:
                    error_count += 1
                    self.stdout.write(
                        self.style.ERROR(f'\nError in row {row_num}: {str(e)}')
                    )
                    self.stdout.write(
                        self.style.ERROR(f'Error type: {type(e).__name__}')
                    )
                    import traceback
                    self.stdout.write(
                        self.style.ERROR(f'Traceback: {traceback.format_exc()}')
                    )
                    continue
            
            # Create remaining records in batch
            if batch:
                Establecimiento.objects.bulk_create(batch, ignore_conflicts=False)
                created_count += len(batch)
            
            self.stdout.write(
                self.style.SUCCESS(
                    f'\n✓ Successfully imported {created_count} records'
                )
            )
            if error_count > 0:
                self.stdout.write(
                    self.style.WARNING(f'{error_count} rows had errors')
                )
                    
        except FileNotFoundError:
            self.stdout.write(
                self.style.ERROR(f'File not found: {excel_file}')
            )
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Unexpected error: {str(e)}')
            )
            import traceback
            self.stdout.write(
                self.style.ERROR(f'Traceback: {traceback.format_exc()}')
            )

    def _clean_row(self, row):
        """Convert Excel row data to appropriate types"""
        
        def safe_get(key):
            """Safely get value from row"""
            val = row.get(key)
            return val if val is not None else ''
        
        def to_int(value):
            try:
                if value is None or value == '':
                    return 0
                value_str = str(value).strip()
                if not value_str:
                    return 0
                return int(float(value_str))
            except (ValueError, TypeError, AttributeError):
                return 0
        
        def to_decimal(value):
            try:
                if value is None or value == '':
                    return Decimal('0.00')
                value_str = str(value).strip()
                if not value_str:
                    return Decimal('0.00')
                return Decimal(value_str)
            except:
                return Decimal('0.00')
        
        def to_string(value):
            try:
                if value is None:
                    return ''
                return str(value).strip()
            except (AttributeError, TypeError):
                return ''
        
        try:
            cleaned = {
                'nombre': to_string(safe_get('nombre')),
                'parroquia': to_string(safe_get('parroquia')),
                'sector': to_string(safe_get('sector')),
                'telefono': to_string(safe_get('telefono')),
                'email': to_string(safe_get('email')) or 'noemail@example.com',
                'propietario': to_string(safe_get('propietario')),
                'tipo_turismo': to_string(safe_get('tipo_turismo')) or 'No especificado',
                'experiencia': to_int(safe_get('experiencia')),
                'asociacion': to_string(safe_get('asociacion')) or 'No',
                'ruc': to_string(safe_get('ruc')) or 'No',
                'estado_local': to_string(safe_get('estado_local')) or 'Propio',
                'servicios_produccion': to_string(safe_get('servicios_produccion')),
                'licencia_gad_loja': to_string(safe_get('licencia_gad_loja')) or 'No',
                'arcsa': to_string(safe_get('arcsa')) or 'No',
                'turismo': to_string(safe_get('turismo')) or 'No',
                'equipos': to_string(safe_get('equipos')),
                'pagina_web': to_string(safe_get('pagina_web')) or 'No',
                'facebook': to_string(safe_get('facebook')) or 'No',
                'instagram': to_string(safe_get('instagram')) or 'No',
                'tiktok': to_string(safe_get('tiktok')) or 'No',
                'whatsapp': to_string(safe_get('whatsapp')) or 'No',
                'tipo': to_string(safe_get('tipo')),
                'mesas': to_int(safe_get('mesas')),
                'plazas': to_int(safe_get('plazas')),
                'banio': to_string(safe_get('banio')) or 'No',
                'complementarios': to_string(safe_get('complementarios')),
                'oferta': to_string(safe_get('oferta')),
                'menu': to_string(safe_get('menu')),
                'tipo_servicio': to_string(safe_get('tipo_servicio')),
                'precio_promedio': to_decimal(safe_get('precio_promedio')),
                'procesos': to_string(safe_get('procesos')),
                'materia_prima': to_string(safe_get('materia_prima')),
                'proveedores': to_string(safe_get('proveedores')),
                'numero_proveedores': to_int(safe_get('numero_proveedores')),
                'numero_mujeres': to_int(safe_get('numero_mujeres')),
                'numero_hombres': to_int(safe_get('numero_hombres')),
                'tiempo_trabajando': to_int(safe_get('tiempo_trabajando')),
                'personal_capacitado': to_string(safe_get('personal_capacitado')) or 'No',
                'frecuencia_capacitacion': to_string(safe_get('frecuencia_capacitacion')),
                'dependencia_ingresos': to_string(safe_get('dependencia_ingresos')),
                'genero': to_string(safe_get('genero')) or 'Otro',
                'nivel_educacion': to_string(safe_get('nivel_educacion')) or 'No especificado',
                'edad': to_int(safe_get('edad')),
                'estado_civil': to_string(safe_get('estado_civil')),
                'longitude': to_decimal(safe_get('longitude')),
                'latitude': to_decimal(safe_get('latitude')),
                'horario': to_string(safe_get('horario')),
                'categoria': to_string(safe_get('categoria')),
                'photo_url': to_string(safe_get('photo_url')) or 'https://picsum.photos/300/30',
                'video_url': to_string(safe_get('video_url')),
                'gallery_urls': self._parse_urls(safe_get('gallery_urls')),
            }
            return cleaned
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Error in _clean_row: {str(e)}')
            )
            raise

    def _parse_urls(self, urls_string):
        """Parse comma-separated URLs into a list"""
        if urls_string is None:
            return []
        urls_str = str(urls_string).strip()
        if not urls_str:
            return []
        return [url.strip() for url in urls_str.split(',') if url.strip()]